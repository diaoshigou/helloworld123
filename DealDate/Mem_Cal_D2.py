
import os
from openpyxl import Workbook
from matplotlib import pyplot as plt
import os
from tkinter import *
import tkinter

wb = Workbook() # 创建文件对象
road = r"E:\Moredian\tools\stability\20230804_172519\com.moredian.mdservice\DumpDevStateService"
ws = wb.active
q = w = a = s = d = f = g = 2
# ws["A1"] = "系统时间"
# ws["B1"] = "argus进程pid"
# ws["C1"] = "CPU占用率"
# ws["D1"] = "apk占用CPU率"
# ws["E1"] = "设备剩余内存"
# ws["F1"] = "apk占用内存"
# ws["G1"] = "设备温度"



def eachFile(filepath):
    pathDir =os.listdir(filepath)        #遍历文件夹中的text
    # print(pathDir)
    return pathDir

def Write(row,data):
    ws[row] = int(data)
    # pass



def quality_D2P(road, eachFile=eachFile):
    # global CPU_radio_list
    System_time_list = []
    Argus_pid_List = []
    CPU_radio_list = []
    apk_cpu_radio_list = []
    Free_RAM_list = []
    apk_RAM_list = []
    Temp_list = []

    global q,w,a,s,d,f,g
    pathDir = eachFile(road)
    for eachDir in pathDir:
        try:

            fopen = open(road + "\\" + eachDir, encoding='UTF-8')  # 设置文件对象
            # with open(road + "\\data.txt", 'a+') as f:
            for lines in fopen.readlines():
                # print(lines)
                # System_time_list_row = "A" + str(q)
                # Argus_pid_List_row = "B" + str(w)
                # CPU_radio_list_row = "C" + str(a)
                # apk_cpu_radio_list_row = "D" + str(s)
                # Free_RAM_list_row = "E" + str(d)
                # apk_RAM_list_row = "F" + str(f)
                # Temp_list_row = "G" + str(g)

                if "camera" in str(lines):
                    System_time = lines[:lines.find("camera") - 1]
                    # ws[System_time_list_row] = str(System_time)
                    System_time_list.append(System_time)
                    q += 1

                if "com.moredian.entrance.guard (pid" in str(lines) and "moredianMem" in str(lines):
                    Argus_pid = lines[lines.find("pid") + 4:lines.find("/") - 1]
                    # Write(Argus_pid_List_row, Argus_pid)
                    w += 1

                if "600%cpu" in str(lines):
                    CPU_radio = (600-int(lines[lines.find("sys") + 5:lines.find("idle")-1]))/6    # D2P

                    # Write(CPU_radio_list_row, CPU_radio)
                    CPU_radio_list.append(float(CPU_radio))
                    a += 1

                if "u0_a15" in str(lines) and "com.moredian.entrance.guard" in str(lines):    # D2PLUS
                    apk_cpu_radio = float(lines[lines.find("com.moredian.entrance.guard") - 21 :lines.find("com.moredian.entrance.guard")-17])
                    # Write(apk_cpu_radio_list_row, apk_cpu_radio)
                    apk_cpu_radio_list.append(int(float(apk_cpu_radio))/6)
                    s += 1

                if "Free RAM" in str(lines):
                    Free_RAM = int(str(lines[lines.find("RAM:") + 7:lines.find("RAM:") +10] + lines[lines.find("K") - 3:lines.find("K") ]))   #D2P
                    # Write(Free_RAM_list_row, Free_RAM)
                    Free_RAM_list.append(int(Free_RAM))
                    d += 1

                if "moredianMem" in str(lines):
                    apk_RAM = int(str(int((lines[lines.find(":") + 1:lines.find(",")])))+str(lines[lines.find(",") + 1:lines.find("K")]))  #D2P
                    # Write(apk_RAM_list_row, apk_RAM)
                    apk_RAM_list.append(int(apk_RAM))
                    f += 1

                if "getTempResult" in str(lines):
                    Temp = lines[lines.find("|") + 7:lines.find("|") + 12]
                    # Write(Temp_list_row, int(Temp) / 1000)
                    Temp_list.append(int(Temp)/1000)
                    g += 1

            fopen.close()

        except Exception as e:
            print(e)
    print("CPU总占用率：" + str(round(sum(CPU_radio_list)/len(CPU_radio_list),2)) + "%(" + str(round(min(CPU_radio_list),2)) +"% - " + str(round(max(CPU_radio_list),2)) + "%)")
    print("apk占用cpu率：" + str(round(sum(apk_cpu_radio_list)/len(apk_cpu_radio_list),2)) + "%(" + str(round(min(apk_cpu_radio_list),2)) +"% - " + str(round(max(apk_cpu_radio_list),2)) + "%)")
    print("系统剩余内存：" + str(round(sum(Free_RAM_list) / len(Free_RAM_list), 2)) + "kb(" + str(
                min(Free_RAM_list)) + "kb - " + str(max(Free_RAM_list)) + "kb)")
    print("apk占用内存：" + str(round(sum(apk_RAM_list)/len(apk_RAM_list),2)) + "kb(" + str(min(apk_RAM_list)) +"kb - " + str(max(apk_RAM_list)) + "kb)")
    print("设备温度：" + str(round(sum(Temp_list)/len(Temp_list),2)) + "℃(" + str(min(Temp_list)) +"℃ - " + str(max(Temp_list)) + "℃)")

def quality_D2C(road, eachFile=eachFile):
    System_time_list = []
    CPU_radio_list = []
    apk_cpu_radio_list = []
    Free_RAM_list = []
    apk_RAM_list = []
    Temp_list = []

    global q,w,a,s,d,f,g
    pathDir = eachFile(road)
    for eachDir in pathDir:
        try:

            fopen = open(road + "\\" + eachDir, encoding='UTF-8')  # 设置文件对象
            for lines in fopen.readlines():

                if "camera" in str(lines):
                    System_time = lines[:lines.find("camera") - 1]
                    # ws[System_time_list_row] = str(System_time)
                    System_time_list.append(System_time)
                    q += 1

                if "com.moredian.entrance.guard (pid" in str(lines) and "moredianMem" in str(lines):
                    Argus_pid = lines[lines.find("pid") + 4:lines.find("/") - 1]
                    # Write(Argus_pid_List_row, Argus_pid)
                    w += 1

                if "600%cpu" in str(lines):
                    CPU_radio = (600-int(lines[lines.find("sys") + 5:lines.find("idle")-1]))/6    # D2P

                    # Write(CPU_radio_list_row, CPU_radio)
                    CPU_radio_list.append(float(CPU_radio))
                    a += 1

                if "u0_a29" in str(lines) and "com.moredian.entrance.guard" in str(lines):    # D2PLUS
                    apk_cpu_radio = float(lines[lines.find("com.moredian.entrance.guard") - 21 :lines.find("com.moredian.entrance.guard")-17])
                    # Write(apk_cpu_radio_list_row, apk_cpu_radio)
                    apk_cpu_radio_list.append(int(float(apk_cpu_radio))/6)
                    s += 1

                if "Free RAM" in str(lines):
                    Free_RAM = int(str(lines[lines.find("RAM:") + 7:lines.find("RAM:") +10] + lines[lines.find("K") - 3:lines.find("K") ]))   #D2P
                    # Write(Free_RAM_list_row, Free_RAM)
                    Free_RAM_list.append(int(Free_RAM))
                    d += 1

                if "moredianMem" in str(lines):
                    apk_RAM = int(str(int((lines[lines.find(":") + 1:lines.find(",")])))+str(lines[lines.find(",") + 1:lines.find("K")]))  #D2P
                    # Write(apk_RAM_list_row, apk_RAM)
                    apk_RAM_list.append(int(apk_RAM))
                    f += 1

                if "getTempResult" in str(lines):
                    Temp = lines[lines.find("|") + 7:lines.find("|") + 12]
                    # Write(Temp_list_row, int(Temp) / 1000)
                    Temp_list.append(int(Temp)/1000)
                    g += 1

            fopen.close()

        except Exception as e:
            print(e)
    print("CPU总占用率：" + str(round(sum(CPU_radio_list)/len(CPU_radio_list),2)) + "%(" + str(round(min(CPU_radio_list),2)) +"% - " + str(round(max(CPU_radio_list),2)) + "%)")
    print("apk占用cpu率：" + str(round(sum(apk_cpu_radio_list)/len(apk_cpu_radio_list),2)) + "%(" + str(round(min(apk_cpu_radio_list),2)) +"% - " + str(round(max(apk_cpu_radio_list),2)) + "%)")
    print("系统剩余内存：" + str(round(sum(Free_RAM_list) / len(Free_RAM_list), 2)) + "kb(" + str(
                min(Free_RAM_list)) + "kb - " + str(max(Free_RAM_list)) + "kb)")
    print("apk占用内存：" + str(round(sum(apk_RAM_list)/len(apk_RAM_list),2)) + "kb(" + str(min(apk_RAM_list)) +"kb - " + str(max(apk_RAM_list)) + "kb)")
    print("设备温度：" + str(round(sum(Temp_list)/len(Temp_list),2)) + "℃(" + str(min(Temp_list)) +"℃ - " + str(max(Temp_list)) + "℃)")

def quality_D2A(road, eachFile=eachFile):
    System_time_list = []
    CPU_radio_list = []
    apk_cpu_radio_list = []
    Free_RAM_list = []
    apk_RAM_list = []
    Temp_list = []

    global q,w,a,s,d,f,g
    pathDir = eachFile(road)
    for eachDir in pathDir:
        try:

            fopen = open(road + "\\" + eachDir, encoding='UTF-8')  # 设置文件对象
            for lines in fopen.readlines():

                if "camera" in str(lines):
                    System_time = lines[:lines.find("camera") - 1]
                    # ws[System_time_list_row] = str(System_time)
                    System_time_list.append(System_time)
                    q += 1

                if "com.moredian.entrance.guard (pid" in str(lines) and "moredianMem" in str(lines):
                    Argus_pid = lines[lines.find("pid") + 4:lines.find("/") - 1]
                    # Write(Argus_pid_List_row, Argus_pid)
                    w += 1

                if "---cpu:U" in str(lines):
                    CPU_radio = (int(lines[lines.find("User") + 5:lines.find("System")-3])+int(lines[lines.find("System") + 7:lines.find("IOW")-3])+int(lines[lines.find("IOW") + 4:lines.find("IRQ")-3]))   # D2P

                    # Write(CPU_radio_list_row, CPU_radio)
                    CPU_radio_list.append(float(CPU_radio))
                    a += 1

                if "u0_a19" in str(lines) and "com.moredian.entrance.guard" in str(lines):    # D2PLUS
                    apk_cpu_radio = float(lines[lines.find("%") - 2 :lines.find("%")])
                    # Write(apk_cpu_radio_list_row, apk_cpu_radio)
                    apk_cpu_radio_list.append(int(apk_cpu_radio))
                    s += 1

                if "Free RAM" in str(lines):
                    Free_RAM = int(str(lines[lines.find("RAM:") + 5:lines.find("k") -1]))
                    # Write(Free_RAM_list_row, Free_RAM)
                    Free_RAM_list.append(int(Free_RAM))
                    d += 1

                if "moredianMem" in str(lines):
                    apk_RAM = int(str(int((lines[lines.find(":") + 1:lines.find("k")-1]))))  #D2P
                    # Write(apk_RAM_list_row, apk_RAM)
                    apk_RAM_list.append(int(apk_RAM))
                    f += 1

                if "getTempResult" in str(lines):
                    Temp = lines[lines.find("|") + 9:lines.find("|") + 14]
                    # Write(Temp_list_row, int(Temp) / 1000)
                    Temp_list.append(int(Temp)/1000)
                    g += 1

            fopen.close()

        except Exception as e:
            print(e)
    print("CPU总占用率：" + str(round(sum(CPU_radio_list)/len(CPU_radio_list),2)) + "%(" + str(round(min(CPU_radio_list),2)) +"% - " + str(round(max(CPU_radio_list),2)) + "%)")
    print("apk占用cpu率：" + str(round(sum(apk_cpu_radio_list)/len(apk_cpu_radio_list),2)) + "%(" + str(round(min(apk_cpu_radio_list),2)) +"% - " + str(round(max(apk_cpu_radio_list),2)) + "%)")
    print("系统剩余内存：" + str(round(sum(Free_RAM_list) / len(Free_RAM_list), 2)) + "kb(" + str(
                min(Free_RAM_list)) + "kb - " + str(max(Free_RAM_list)) + "kb)")
    print("apk占用内存：" + str(round(sum(apk_RAM_list)/len(apk_RAM_list),2)) + "kb(" + str(min(apk_RAM_list)) +"kb - " + str(max(apk_RAM_list)) + "kb)")
    print("设备温度：" + str(round(sum(Temp_list)/len(Temp_list),2)) + "℃(" + str(min(Temp_list)) +"℃ - " + str(max(Temp_list)) + "℃)")

def quality_D3(road, eachFile=eachFile):
    System_time_list = []
    CPU_radio_list = []
    apk_cpu_radio_list = []
    Free_RAM_list = []
    apk_RAM_list = []
    Temp_list = []

    global q,w,a,s,d,f,g
    pathDir = eachFile(road)
    for eachDir in pathDir:
        try:

            fopen = open(road + "\\" + eachDir, encoding='UTF-8')  # 设置文件对象
            for lines in fopen.readlines():

                if "camera" in str(lines):
                    System_time = lines[:lines.find("camera") - 1]
                    # ws[System_time_list_row] = str(System_time)
                    System_time_list.append(System_time)
                    q += 1

                if "com.moredian.entrance.guard (pid" in str(lines) and "moredianMem" in str(lines):
                    Argus_pid = lines[lines.find("pid") + 4:lines.find("/") - 1]
                    # Write(Argus_pid_List_row, Argus_pid)
                    w += 1

                if "800%" in str(lines):
                    CPU_radio = (800 - int(lines[lines.find("sys") + 4:lines.find("idle")-1]))   # D2P

                    # Write(CPU_radio_list_row, CPU_radio)
                    CPU_radio_list.append(float(CPU_radio)/8)
                    a += 1

                if "u0_a16" in str(lines) and "com.moredian.entrance.guard" in str(lines):    # D2PLUS
                    apk_cpu_radio = float(lines[lines.find("M") + 10 :lines.find("com.moredian.entrance.guard")-17])
                    # Write(apk_cpu_radio_list_row, apk_cpu_radio)
                    apk_cpu_radio_list.append(int(apk_cpu_radio)/8)
                    s += 1

                if "Free RAM" in str(lines):
                    Free_RAM = int(str(lines[lines.find("RAM:") + 6:lines.find(",")])+str(lines[lines.find(",") + 1:lines.find("K")]))
                    # Write(Free_RAM_list_row, Free_RAM)
                    Free_RAM_list.append(int(Free_RAM))
                    d += 1

                if "moredianMem" in str(lines):
                    apk_RAM = int(str(int((lines[lines.find(":") + 1:lines.find(",")])+str(lines[lines.find(",") + 1:lines.find("K")]))))  #D2P
                    # Write(apk_RAM_list_row, apk_RAM)
                    apk_RAM_list.append(int(apk_RAM))
                    f += 1

                if "getTempResult" in str(lines):
                    Temp = lines[lines.find("|") + 9:lines.find("|") + 14]
                    # Write(Temp_list_row, int(Temp) / 1000)
                    Temp_list.append(int(Temp)/1000)
                    g += 1

            fopen.close()

        except Exception as e:
            print(e)
    CPU_radio = "CPU总占用率：" + str(round(sum(CPU_radio_list)/len(CPU_radio_list),2)) + "%(" + str(round(min(CPU_radio_list),2)) +"% - " + str(round(max(CPU_radio_list),2)) + "%)"
    apk_cpu_radio = "apk占用cpu率：" + str(round(sum(apk_cpu_radio_list)/len(apk_cpu_radio_list),2)) + "%(" + str(round(min(apk_cpu_radio_list),2)) +"% - " + str(round(max(apk_cpu_radio_list),2)) + "%)"
    # print("CPU总占用率：" + str(round(sum(CPU_radio_list)/len(CPU_radio_list),2)) + "%(" + str(round(min(CPU_radio_list),2)) +"% - " + str(round(max(CPU_radio_list),2)) + "%)")
    # print("apk占用cpu率：" + str(round(sum(apk_cpu_radio_list)/len(apk_cpu_radio_list),2)) + "%(" + str(round(min(apk_cpu_radio_list),2)) +"% - " + str(round(max(apk_cpu_radio_list),2)) + "%)")
    # print("系统剩余内存：" + str(round(sum(Free_RAM_list) / len(Free_RAM_list), 2)) + "kb(" + str(
    #             min(Free_RAM_list)) + "kb - " + str(max(Free_RAM_list)) + "kb)")
    # print("apk占用内存：" + str(round(sum(apk_RAM_list)/len(apk_RAM_list),2)) + "kb(" + str(min(apk_RAM_list)) +"kb - " + str(max(apk_RAM_list)) + "kb)")
    # print("设备温度：" + str(round(sum(Temp_list)/len(Temp_list),2)) + "℃(" + str(min(Temp_list)) +"℃ - " + str(max(Temp_list)) + "℃)")
    tuple = (CPU_radio,apk_cpu_radio)
    return tuple

def finddata():
    lb.delete(0, END)
    path = filepath.get()
    filePath = path
    tuple = quality_D3(road)
    lb.insert(END,str(tuple[0]) )




# quality_D3(road)
# quality_D2A(road)
# quality_D2C(road)
# quality_D2P(road)
# wb.save(r"C:\Users\86183\Desktop\logs\mdService\files\logs\a.xlsx")
root = Tk()
root.title("Quality")
root.geometry('750x350') # 全局界面大小
root.wm_attributes('-topmost', 1) # 永远保持最前
sb = Scrollbar(root)    #垂直滑动条
sbx =Scrollbar(root,orient='horizontal')    # 水平滑动条
listbox = StringVar()
filepath = StringVar()




DealFile_Path = tkinter.Checkbutton(root,text="D3",variable = quality_D3(road, eachFile=eachFile),borderwidth=1,width=5,anchor=W).place(x=150, y=50, anchor="se")
lb = Listbox(root,yscrollcommand= sb.set,xscrollcommand=sbx.set,width=50,listvariable=listbox,selectmode='extended')
Data_FilePath = tkinter.Label(root,text="文件路径：",borderwidth=5,width=15,anchor=NE).grid(row=2,column=2)
Entry_FilePath = Entry(root, borderwidth=1,textvariable=filepath,width=40).grid(row=2,column=3,columnspan=2)
comand = Button(root, text="获取", command=finddata,width=10, height=2,borderwidth=2).grid(row=2, column=6,rowspan=3,sticky=W)

# sb.grid(row=5, column=10,sticky='ENS',rowspan=5)
# sbx.grid(row=10,column=3,columnspan=7,sticky='SWE')
sb.place(x=450, y=60)

# lb.grid(row=5,sticky='W', column=3,rowspan=5,columnspan=7,ipadx=100)
lb.place(x=100, y=50)
sb.config(command=lb.yview)
sbx.config(command=lb.xview)
root.mainloop()


root.mainloop()